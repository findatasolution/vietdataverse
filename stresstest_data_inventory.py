#!/usr/bin/env python3
"""
Viet Dataverse — Data Inventory Generator
Generates stresstest_data_inventory.xlsx with:
  - Sheet "System DB Inventory"   : all tables + metadata
  - Sheet "Data Dictionary"       : all columns + descriptions (mandatory fields auto-generated)
  - Sheet "Stresstest Inventory"  : crawl tracking status

Mandatory columns for every table:
  id, period, crawl_time, source, group_name

Data Group Taxonomy (max 5):
  macro     — VN macro indicators (CPI, GDP, trade, IIP, PPI)
  finance   — Financial rates (term deposit, SBV interbank, FX)
  commodity — Commodity prices (gold, silver, oil, global)
  stock     — Equity & corporate (VN30 OHLCV, financials, ratios)
  sentiment — Market sentiment [future]

Usage:
  python stresstest_data_inventory.py
Output:
  productowner/stresstest_data_inventory.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "productowner", "stresstest_data_inventory.xlsx")

# ─────────────────────────────────────────────
# GROUP TAXONOMY
# ─────────────────────────────────────────────

GROUPS = {
    "macro":     "Vietnam macro indicators (CPI, GDP, trade, IIP, PPI)",
    "finance":   "Financial market rates (term deposit, SBV interbank, FX)",
    "commodity": "Commodity prices (gold, silver, oil, global)",
    "stock":     "Equity & corporate (VN30 OHLCV, financials, ratios)",
    "sentiment": "Market sentiment indicators [future: news, social, retail flow]",
}

GROUP_COLORS = {
    "macro":     "D9E1F2",  # blue
    "finance":   "E2EFDA",  # green
    "commodity": "FFF2CC",  # yellow
    "stock":     "FCE4D6",  # orange
    "sentiment": "EAD1DC",  # pink
}

# ─────────────────────────────────────────────
# DATA: System DB Inventory
# ─────────────────────────────────────────────

DB_INVENTORY = [
    # ── CRAWLING_BOT_DB — Finance ─────────────────────────
    {
        "db": "CRAWLING_BOT_DB",
        "table": "vn_macro_termdepo_daily",
        "group": "finance",
        "crawler": "crawl_bank_termdepo.py",
        "entity_key": "bank_code, date",
        "period_col": "date (DATE)",
        "freq": "daily",
        "source_org": "ACB Bank",
        "source_url": "acb.com.vn",
        "description": "Lãi suất tiền gửi có kỳ hạn ngân hàng ACB (1m–36m)",
        "from_year": 2020,
        "rows_approx": 2500,
        "active": True,
    },
    {
        "db": "CRAWLING_BOT_DB",
        "table": "vn_macro_fxrate_daily",
        "group": "finance",
        "crawler": "crawl_exchange_rate.py",
        "entity_key": "type, source, bank, date",
        "period_col": "date (DATE)",
        "freq": "daily",
        "source_org": "SBV, VNAppMob",
        "source_url": "sbv.gov.vn / api.vnappmob.com",
        "description": "Tỷ giá hối đoái: USD, EUR, JPY, GBP, CNY... (VND/ngoại tệ)",
        "from_year": 2019,
        "rows_approx": 30000,
        "active": True,
    },
    {
        "db": "CRAWLING_BOT_DB",
        "table": "vn_macro_sbv_rate_daily",
        "group": "finance",
        "crawler": "crawl_sbv.py",
        "entity_key": "date",
        "period_col": "date (DATE)",
        "freq": "daily",
        "source_org": "SBV",
        "source_url": "sbv.gov.vn",
        "description": "Lãi suất liên ngân hàng SBV: qua đêm, 1W–9M + lãi suất cơ bản",
        "from_year": 2020,
        "rows_approx": 1500,
        "active": True,
    },
    # ── CRAWLING_BOT_DB — Commodity ───────────────────────
    {
        "db": "CRAWLING_BOT_DB",
        "table": "vn_macro_gold_daily",
        "group": "commodity",
        "crawler": "crawl_gold_silver.py",
        "entity_key": "type, date",
        "period_col": "date (DATE)",
        "freq": "twice_daily",
        "source_org": "24h.com.vn, BTMC",
        "source_url": "24h.com.vn / api.btmc.vn",
        "description": "Giá vàng VN: SJC, DOJI, PNJ, BTMC (mua/bán, VND/chỉ)",
        "from_year": 2018,
        "rows_approx": 50000,
        "active": True,
    },
    {
        "db": "CRAWLING_BOT_DB",
        "table": "vn_macro_silver_daily",
        "group": "commodity",
        "crawler": "crawl_gold_silver.py",
        "entity_key": "date",
        "period_col": "date (DATE)",
        "freq": "twice_daily",
        "source_org": "Phú Quý",
        "source_url": "phuquy.com.vn",
        "description": "Giá bạc Phú Quý (mua/bán, VND)",
        "from_year": 2020,
        "rows_approx": 5000,
        "active": True,
    },
    # ── CRAWLING_BOT_DB — Macro ───────────────────────────
    {
        "db": "CRAWLING_BOT_DB",
        "table": "vn_gso_cpi_monthly",
        "group": "macro",
        "crawler": "crawl_gso_cpi.py",
        "entity_key": "period, category",
        "period_col": "period (VARCHAR 7, YYYY-MM)",
        "freq": "monthly",
        "source_org": "GSO / NSO",
        "source_url": "gso.gov.vn / nso.gov.vn",
        "description": "CPI Việt Nam hàng tháng: tổng + theo danh mục (thực phẩm, nhà ở...)",
        "from_year": 2000,
        "rows_approx": 3000,
        "active": True,
    },
    {
        "db": "CRAWLING_BOT_DB",
        "table": "vn_gso_ppi_monthly",
        "group": "macro",
        "crawler": "⚠️ Chưa có crawler",
        "entity_key": "period, sector",
        "period_col": "period (VARCHAR 7, YYYY-MM)",
        "freq": "monthly",
        "source_org": "GSO",
        "source_url": "gso.gov.vn",
        "description": "Chỉ số giá sản xuất (PPI) theo ngành",
        "from_year": 2010,
        "rows_approx": 1500,
        "active": False,
    },
    {
        "db": "CRAWLING_BOT_DB",
        "table": "vn_gso_gdp_quarterly",
        "group": "macro",
        "crawler": "crawl_gso_gdp.py",
        "entity_key": "year, quarter, sector",
        "period_col": "year + quarter (INT)",
        "freq": "quarterly",
        "source_org": "GSO",
        "source_url": "gso.gov.vn",
        "description": "GDP Việt Nam hàng quý theo ngành (nông nghiệp, công nghiệp, dịch vụ)",
        "from_year": 2010,
        "rows_approx": 800,
        "active": False,
    },
    {
        "db": "CRAWLING_BOT_DB",
        "table": "vn_gso_trade_monthly",
        "group": "macro",
        "crawler": "crawl_gso_trade.py",
        "entity_key": "period",
        "period_col": "period (VARCHAR 7, YYYY-MM)",
        "freq": "monthly",
        "source_org": "GSO / Hải quan",
        "source_url": "gso.gov.vn / customs.gov.vn",
        "description": "Xuất nhập khẩu Việt Nam hàng tháng (tỷ USD)",
        "from_year": 2010,
        "rows_approx": 300,
        "active": False,
    },
    {
        "db": "CRAWLING_BOT_DB",
        "table": "vn_gso_iip_monthly",
        "group": "macro",
        "crawler": "crawl_gso_industry.py",
        "entity_key": "period, sector_name",
        "period_col": "period (VARCHAR 7, YYYY-MM)",
        "freq": "monthly",
        "source_org": "GSO",
        "source_url": "gso.gov.vn",
        "description": "Chỉ số Sản xuất Công nghiệp (IIP) theo ngành hàng tháng",
        "from_year": 2015,
        "rows_approx": 2000,
        "active": False,
    },
    # ── GLOBAL_INDICATOR_DB — Commodity ───────────────────
    {
        "db": "GLOBAL_INDICATOR_DB",
        "table": "global_macro",
        "group": "commodity",
        "crawler": "crawl_gold_silver.py",
        "entity_key": "date",
        "period_col": "date (DATE)",
        "freq": "daily",
        "source_org": "Yahoo Finance",
        "source_url": "Yahoo Finance (yfinance)",
        "description": "Giá vàng quốc tế (GC=F), bạc (SI=F), NASDAQ (^IXIC) — USD",
        "from_year": 2020,
        "rows_approx": 1500,
        "active": True,
    },
    # ── CRAWLING_CORP_DB — Stock ──────────────────────────
    {
        "db": "CRAWLING_CORP_DB",
        "table": "vn30_ohlcv_daily",
        "group": "stock",
        "crawler": "crawl_vn30_price.py",
        "entity_key": "ticker, date",
        "period_col": "date (DATE)",
        "freq": "daily",
        "source_org": "vnstock3 (SSI/TCBS)",
        "source_url": "vnstock3 API",
        "description": "Giá OHLCV hàng ngày cho 30 cổ phiếu VN30 (HSX)",
        "from_year": 2018,
        "rows_approx": 150000,
        "active": True,
    },
    {
        "db": "CRAWLING_CORP_DB",
        "table": "vn30_company_profile",
        "group": "stock",
        "crawler": "crawl_vn30_profile.py",
        "entity_key": "ticker",
        "period_col": "updated_at (TIMESTAMP)",
        "freq": "monthly",
        "source_org": "vnstock3",
        "source_url": "vnstock3 API",
        "description": "Hồ sơ công ty VN30: tên, ngành ICB, vốn hóa, ngày niêm yết",
        "from_year": 2024,
        "rows_approx": 30,
        "active": True,
    },
    {
        "db": "CRAWLING_CORP_DB",
        "table": "vn30_income_stmt_quarterly",
        "group": "stock",
        "crawler": "crawl_vn30_financials.py",
        "entity_key": "ticker, year, quarter",
        "period_col": "year + quarter (INT)",
        "freq": "quarterly",
        "source_org": "vnstock3",
        "source_url": "vnstock3 API",
        "description": "Kết quả kinh doanh hàng quý: doanh thu, lợi nhuận, EPS",
        "from_year": 2019,
        "rows_approx": 2400,
        "active": True,
    },
    {
        "db": "CRAWLING_CORP_DB",
        "table": "vn30_balance_sheet_quarterly",
        "group": "stock",
        "crawler": "crawl_vn30_financials.py",
        "entity_key": "ticker, year, quarter",
        "period_col": "year + quarter (INT)",
        "freq": "quarterly",
        "source_org": "vnstock3",
        "source_url": "vnstock3 API",
        "description": "Bảng cân đối kế toán hàng quý: tổng tài sản, nợ, vốn chủ sở hữu",
        "from_year": 2019,
        "rows_approx": 2400,
        "active": True,
    },
    {
        "db": "CRAWLING_CORP_DB",
        "table": "vn30_cashflow_quarterly",
        "group": "stock",
        "crawler": "crawl_vn30_financials.py",
        "entity_key": "ticker, year, quarter",
        "period_col": "year + quarter (INT)",
        "freq": "quarterly",
        "source_org": "vnstock3",
        "source_url": "vnstock3 API",
        "description": "Lưu chuyển tiền tệ hàng quý: hoạt động, đầu tư, tài chính",
        "from_year": 2019,
        "rows_approx": 2400,
        "active": True,
    },
    {
        "db": "CRAWLING_CORP_DB",
        "table": "vn30_ratio_daily",
        "group": "stock",
        "crawler": "crawl_vn30_ratios.py",
        "entity_key": "ticker, date",
        "period_col": "date (DATE)",
        "freq": "daily",
        "source_org": "vnstock3",
        "source_url": "vnstock3 API",
        "description": "Chỉ số tài chính hàng ngày: P/E, P/B, ROE, ROA, EPS, Dividend Yield",
        "from_year": 2020,
        "rows_approx": 50000,
        "active": True,
    },
]

# ─────────────────────────────────────────────
# DATA: Data Dictionary — table-specific columns
# (Mandatory fields id, period, crawl_time, source, group_name auto-generated per table)
# ─────────────────────────────────────────────

TABLE_SPECIFIC_COLS = {
    "vn_macro_termdepo_daily": [
        {"column": "bank_code",  "type": "VARCHAR(10)", "description": "Mã ngân hàng",                     "unit": "—",        "example": "ACB"},
        {"column": "term_1m",    "type": "FLOAT",       "description": "Lãi suất kỳ hạn 1 tháng",          "unit": "% năm",    "example": "3.2"},
        {"column": "term_3m",    "type": "FLOAT",       "description": "Lãi suất kỳ hạn 3 tháng",          "unit": "% năm",    "example": "4.1"},
        {"column": "term_6m",    "type": "FLOAT",       "description": "Lãi suất kỳ hạn 6 tháng",          "unit": "% năm",    "example": "4.8"},
        {"column": "term_9m",    "type": "FLOAT",       "description": "Lãi suất kỳ hạn 9 tháng",          "unit": "% năm",    "example": "5.0"},
        {"column": "term_12m",   "type": "FLOAT",       "description": "Lãi suất kỳ hạn 12 tháng",         "unit": "% năm",    "example": "5.5"},
        {"column": "term_24m",   "type": "FLOAT",       "description": "Lãi suất kỳ hạn 24 tháng",         "unit": "% năm",    "example": "5.8"},
        {"column": "term_36m",   "type": "FLOAT",       "description": "Lãi suất kỳ hạn 36 tháng",         "unit": "% năm",    "example": "6.0"},
    ],
    "vn_macro_fxrate_daily": [
        {"column": "type",         "type": "VARCHAR(20)", "description": "Mã ngoại tệ (ISO 4217)",           "unit": "—",              "example": "USD, EUR, JPY"},
        {"column": "bank",         "type": "VARCHAR(10)", "description": "Ngân hàng báo giá",                "unit": "—",              "example": "SBV, BID, TCB"},
        {"column": "buy_cash",     "type": "FLOAT",       "description": "Tỷ giá mua tiền mặt",             "unit": "VND/ngoại tệ",   "example": "25380.0"},
        {"column": "buy_transfer", "type": "FLOAT",       "description": "Tỷ giá mua chuyển khoản",         "unit": "VND/ngoại tệ",   "example": "25390.0"},
        {"column": "sell_rate",    "type": "FLOAT",       "description": "Tỷ giá bán",                      "unit": "VND/ngoại tệ",   "example": "25650.0"},
    ],
    "vn_macro_sbv_rate_daily": [
        {"column": "quadem",           "type": "FLOAT", "description": "Lãi suất qua đêm liên ngân hàng",  "unit": "% năm", "example": "4.2"},
        {"column": "w1",               "type": "FLOAT", "description": "Lãi suất kỳ hạn 1 tuần",          "unit": "% năm", "example": "4.3"},
        {"column": "m1",               "type": "FLOAT", "description": "Lãi suất kỳ hạn 1 tháng",         "unit": "% năm", "example": "4.5"},
        {"column": "m3",               "type": "FLOAT", "description": "Lãi suất kỳ hạn 3 tháng",         "unit": "% năm", "example": "4.6"},
        {"column": "m6",               "type": "FLOAT", "description": "Lãi suất kỳ hạn 6 tháng",         "unit": "% năm", "example": "4.7"},
        {"column": "refinancing_rate", "type": "FLOAT", "description": "Lãi suất tái cấp vốn SBV",        "unit": "% năm", "example": "4.5"},
        {"column": "rediscount_rate",  "type": "FLOAT", "description": "Lãi suất tái chiết khấu SBV",     "unit": "% năm", "example": "3.0"},
    ],
    "vn_macro_gold_daily": [
        {"column": "type",       "type": "VARCHAR(100)", "description": "Loại vàng (nguồn + loại sản phẩm)", "unit": "—",       "example": "SJC HN, BTMC SJC"},
        {"column": "buy_price",  "type": "FLOAT",        "description": "Giá mua vào",                       "unit": "VND/chỉ", "example": "85500000"},
        {"column": "sell_price", "type": "FLOAT",        "description": "Giá bán ra",                        "unit": "VND/chỉ", "example": "87000000"},
    ],
    "vn_macro_silver_daily": [
        {"column": "buy_price",  "type": "FLOAT", "description": "Giá bạc mua vào", "unit": "VND", "example": "990000"},
        {"column": "sell_price", "type": "FLOAT", "description": "Giá bạc bán ra",  "unit": "VND", "example": "1010000"},
    ],
    "vn_gso_cpi_monthly": [
        {"column": "category",    "type": "VARCHAR(100)", "description": "Danh mục CPI (tổng, thực phẩm, nhà ở...)", "unit": "—",     "example": "overall, food"},
        {"column": "cpi_index",   "type": "FLOAT",        "description": "Chỉ số CPI (gốc = 100)",                    "unit": "điểm",  "example": "112.5"},
        {"column": "cpi_mom_pct", "type": "FLOAT",        "description": "Thay đổi CPI so tháng trước (MoM)",         "unit": "%",     "example": "0.23"},
        {"column": "cpi_yoy_pct", "type": "FLOAT",        "description": "Thay đổi CPI so cùng kỳ năm trước (YoY, compound 12m)", "unit": "%", "example": "3.14"},
    ],
    "vn_gso_ppi_monthly": [
        {"column": "sector",      "type": "VARCHAR(100)", "description": "Ngành sản xuất",                   "unit": "—",    "example": "Manufacturing"},
        {"column": "ppi_index",   "type": "FLOAT",        "description": "Chỉ số PPI (gốc = 100)",           "unit": "điểm", "example": "105.3"},
        {"column": "ppi_mom_pct", "type": "FLOAT",        "description": "Thay đổi PPI MoM",                 "unit": "%",    "example": "0.5"},
        {"column": "ppi_yoy_pct", "type": "FLOAT",        "description": "Thay đổi PPI YoY",                 "unit": "%",    "example": "4.2"},
    ],
    "vn_gso_gdp_quarterly": [
        {"column": "year",            "type": "INTEGER", "description": "Năm báo cáo",                      "unit": "—",            "example": "2024"},
        {"column": "quarter",         "type": "INTEGER", "description": "Quý (1-4)",                         "unit": "—",            "example": "3"},
        {"column": "sector",          "type": "VARCHAR(50)", "description": "Ngành kinh tế",                 "unit": "—",            "example": "total, services"},
        {"column": "gdp_billion_vnd", "type": "FLOAT",   "description": "GDP danh nghĩa",                   "unit": "tỷ VND",       "example": "2350000.0"},
        {"column": "growth_yoy_pct",  "type": "FLOAT",   "description": "Tăng trưởng GDP so cùng kỳ",       "unit": "%",            "example": "6.8"},
    ],
    "vn_gso_trade_monthly": [
        {"column": "export_billion_usd", "type": "FLOAT", "description": "Giá trị xuất khẩu",              "unit": "tỷ USD", "example": "35.2"},
        {"column": "import_billion_usd", "type": "FLOAT", "description": "Giá trị nhập khẩu",              "unit": "tỷ USD", "example": "33.1"},
        {"column": "trade_balance",      "type": "FLOAT", "description": "Cán cân thương mại (XK - NK)",    "unit": "tỷ USD", "example": "2.1"},
        {"column": "yoy_export_pct",     "type": "FLOAT", "description": "Tăng trưởng xuất khẩu YoY",      "unit": "%",       "example": "12.5"},
        {"column": "yoy_import_pct",     "type": "FLOAT", "description": "Tăng trưởng nhập khẩu YoY",      "unit": "%",       "example": "8.3"},
    ],
    "vn_gso_iip_monthly": [
        {"column": "sector_name", "type": "VARCHAR(200)", "description": "Tên ngành sản xuất",              "unit": "—",    "example": "Điện, điện tử"},
        {"column": "iip_index",   "type": "FLOAT",        "description": "Chỉ số IIP (gốc = 100)",          "unit": "điểm", "example": "118.4"},
        {"column": "iip_yoy_pct", "type": "FLOAT",        "description": "Tăng trưởng IIP so cùng kỳ",     "unit": "%",    "example": "8.2"},
    ],
    "vn30_ohlcv_daily": [
        {"column": "ticker", "type": "VARCHAR(10)", "description": "Mã chứng khoán VN30",                   "unit": "—",    "example": "VCB, BID, CTG"},
        {"column": "open",   "type": "FLOAT",       "description": "Giá mở cửa phiên",                      "unit": "VND",  "example": "82000.0"},
        {"column": "high",   "type": "FLOAT",       "description": "Giá cao nhất trong phiên",              "unit": "VND",  "example": "84000.0"},
        {"column": "low",    "type": "FLOAT",       "description": "Giá thấp nhất trong phiên",             "unit": "VND",  "example": "81500.0"},
        {"column": "close",  "type": "FLOAT",       "description": "Giá đóng cửa phiên",                    "unit": "VND",  "example": "83200.0"},
        {"column": "volume", "type": "FLOAT",       "description": "Khối lượng giao dịch",                  "unit": "cổ phiếu", "example": "5200000.0"},
        {"column": "value",  "type": "FLOAT",       "description": "Giá trị giao dịch",                     "unit": "VND",  "example": "432640000000"},
    ],
    "vn30_company_profile": [
        {"column": "ticker",             "type": "VARCHAR(10)",  "description": "Mã chứng khoán",            "unit": "—",       "example": "VCB"},
        {"column": "company_name",       "type": "VARCHAR(200)", "description": "Tên công ty tiếng Việt",    "unit": "—",       "example": "Vietcombank"},
        {"column": "exchange",           "type": "VARCHAR(10)",  "description": "Sàn giao dịch",             "unit": "—",       "example": "HOSE"},
        {"column": "icb_sector",         "type": "VARCHAR(100)", "description": "Ngành ICB cấp cao nhất",    "unit": "—",       "example": "Financials"},
        {"column": "icb_industry",       "type": "VARCHAR(100)", "description": "Ngành ICB cấp 2",           "unit": "—",       "example": "Banking"},
        {"column": "market_cap_billion", "type": "FLOAT",        "description": "Vốn hóa thị trường",        "unit": "tỷ VND",  "example": "520000.0"},
        {"column": "listed_date",        "type": "DATE",         "description": "Ngày niêm yết",             "unit": "YYYY-MM-DD", "example": "2009-06-22"},
    ],
    "vn30_income_stmt_quarterly": [
        {"column": "ticker",       "type": "VARCHAR(10)", "description": "Mã chứng khoán",                  "unit": "—",      "example": "VCB"},
        {"column": "year",         "type": "INTEGER",     "description": "Năm tài chính",                    "unit": "—",      "example": "2024"},
        {"column": "quarter",      "type": "INTEGER",     "description": "Quý (1-4)",                        "unit": "—",      "example": "3"},
        {"column": "revenue",      "type": "FLOAT",       "description": "Doanh thu thuần",                  "unit": "tỷ VND", "example": "25000.0"},
        {"column": "gross_profit", "type": "FLOAT",       "description": "Lợi nhuận gộp",                   "unit": "tỷ VND", "example": "18000.0"},
        {"column": "net_income",   "type": "FLOAT",       "description": "Lợi nhuận sau thuế",               "unit": "tỷ VND", "example": "12000.0"},
        {"column": "eps",          "type": "FLOAT",       "description": "Thu nhập trên mỗi cổ phiếu (EPS)", "unit": "VND",    "example": "5200.0"},
    ],
    "vn30_balance_sheet_quarterly": [
        {"column": "ticker",            "type": "VARCHAR(10)", "description": "Mã chứng khoán",             "unit": "—",       "example": "VCB"},
        {"column": "year",              "type": "INTEGER",     "description": "Năm tài chính",               "unit": "—",       "example": "2024"},
        {"column": "quarter",           "type": "INTEGER",     "description": "Quý (1-4)",                   "unit": "—",       "example": "3"},
        {"column": "total_assets",      "type": "FLOAT",       "description": "Tổng tài sản",                "unit": "tỷ VND",  "example": "1900000.0"},
        {"column": "total_liabilities", "type": "FLOAT",       "description": "Tổng nợ phải trả",            "unit": "tỷ VND",  "example": "1750000.0"},
        {"column": "equity",            "type": "FLOAT",       "description": "Vốn chủ sở hữu",             "unit": "tỷ VND",  "example": "150000.0"},
        {"column": "cash",              "type": "FLOAT",       "description": "Tiền và tương đương tiền",    "unit": "tỷ VND",  "example": "35000.0"},
    ],
    "vn30_cashflow_quarterly": [
        {"column": "ticker",        "type": "VARCHAR(10)", "description": "Mã chứng khoán",                 "unit": "—",      "example": "VCB"},
        {"column": "year",          "type": "INTEGER",     "description": "Năm tài chính",                   "unit": "—",      "example": "2024"},
        {"column": "quarter",       "type": "INTEGER",     "description": "Quý (1-4)",                       "unit": "—",      "example": "3"},
        {"column": "cfo",           "type": "FLOAT",       "description": "Lưu chuyển từ hoạt động kinh doanh", "unit": "tỷ VND", "example": "15000.0"},
        {"column": "cfi",           "type": "FLOAT",       "description": "Lưu chuyển từ đầu tư",           "unit": "tỷ VND", "example": "-5000.0"},
        {"column": "cff",           "type": "FLOAT",       "description": "Lưu chuyển từ tài chính",        "unit": "tỷ VND", "example": "-3000.0"},
        {"column": "free_cashflow", "type": "FLOAT",       "description": "Dòng tiền tự do (CFO - CapEx)",   "unit": "tỷ VND", "example": "12000.0"},
    ],
    "vn30_ratio_daily": [
        {"column": "ticker",             "type": "VARCHAR(10)", "description": "Mã chứng khoán",            "unit": "—",      "example": "VCB"},
        {"column": "pe",                 "type": "FLOAT",       "description": "P/E — Giá/Thu nhập",        "unit": "lần",    "example": "12.5"},
        {"column": "pb",                 "type": "FLOAT",       "description": "P/B — Giá/Sổ sách",        "unit": "lần",    "example": "1.8"},
        {"column": "ps",                 "type": "FLOAT",       "description": "P/S — Giá/Doanh thu",      "unit": "lần",    "example": "2.1"},
        {"column": "roe",                "type": "FLOAT",       "description": "Tỷ suất sinh lời vốn chủ", "unit": "%",      "example": "18.3"},
        {"column": "roa",                "type": "FLOAT",       "description": "Tỷ suất sinh lời tài sản", "unit": "%",      "example": "1.8"},
        {"column": "eps",                "type": "FLOAT",       "description": "Thu nhập trên cổ phiếu",   "unit": "VND",    "example": "5200.0"},
        {"column": "dividend_yield",     "type": "FLOAT",       "description": "Tỷ suất cổ tức",           "unit": "%",      "example": "2.5"},
        {"column": "market_cap_billion", "type": "FLOAT",       "description": "Vốn hóa thị trường",       "unit": "tỷ VND", "example": "520000.0"},
    ],
    "global_macro": [
        {"column": "gold_price",   "type": "FLOAT", "description": "Giá vàng quốc tế (GC=F)",          "unit": "USD/oz",  "example": "2350.5"},
        {"column": "silver_price", "type": "FLOAT", "description": "Giá bạc quốc tế (SI=F)",           "unit": "USD/oz",  "example": "28.4"},
        {"column": "nasdaq_price", "type": "FLOAT", "description": "Chỉ số NASDAQ Composite (^IXIC)",  "unit": "điểm",   "example": "17800.0"},
    ],
}

# ─────────────────────────────────────────────
# Build full Data Dictionary with mandatory cols
# ─────────────────────────────────────────────

def build_full_dictionary():
    """
    Auto-prepend mandatory columns (id, period, crawl_time, source, group_name)
    for every table, then append table-specific columns.
    """
    rows = []
    for tbl in DB_INVENTORY:
        table    = tbl["table"]
        group    = tbl["group"]
        src_url  = tbl["source_url"]
        period   = tbl["period_col"]

        mandatory = [
            {"table": table, "mandatory": "✅", "column": "id",         "type": "SERIAL / INTEGER PK", "description": "Primary key, auto-increment (SERIAL)",        "unit": "—",      "example": "1"},
            {"table": table, "mandatory": "✅", "column": "period",      "type": period,                "description": "Thời điểm dữ liệu: DATE (daily) hoặc VARCHAR(7) YYYY-MM (monthly)", "unit": "—", "example": "2025-01-15 / 2025-01"},
            {"table": table, "mandatory": "✅", "column": "crawl_time",  "type": "TIMESTAMP NOT NULL",  "description": "Thời điểm crawl (UTC)",                       "unit": "UTC",    "example": "2025-01-15 02:30:00"},
            {"table": table, "mandatory": "✅", "column": "source",      "type": "TEXT",                "description": f"URL hoặc tên tổ chức nguồn ({src_url})",      "unit": "—",      "example": src_url.split("/")[0]},
            {"table": table, "mandatory": "✅", "column": "group_name",  "type": "VARCHAR(20)",         "description": f"Nhóm dữ liệu: '{group}' — {GROUPS[group]}", "unit": "—",      "example": group},
        ]

        specific = [
            {"table": table, "mandatory": "—", **col}
            for col in TABLE_SPECIFIC_COLS.get(table, [])
        ]

        rows.extend(mandatory)
        rows.extend(specific)
    return rows

# ─────────────────────────────────────────────
# DATA: Stresstest Tracking
# ─────────────────────────────────────────────

STRESSTEST_TRACKING = [
    {"category": "finance",   "indicator": "Term Deposit ACB (1m-36m)",            "table": "vn_macro_termdepo_daily",       "has_data": "Có",   "from_year": 2020, "notes": "ACB only — daily"},
    {"category": "finance",   "indicator": "Lãi suất liên ngân hàng SBV",          "table": "vn_macro_sbv_rate_daily",       "has_data": "Có",   "from_year": 2020, "notes": "Overnight, 1W-9M, policy rates"},
    {"category": "finance",   "indicator": "Tỷ giá USD, EUR, JPY, GBP...",         "table": "vn_macro_fxrate_daily",         "has_data": "Có",   "from_year": 2019, "notes": "SBV central + bank commercial rates"},
    {"category": "commodity", "indicator": "Giá vàng SJC, DOJI, PNJ, BTMC",       "table": "vn_macro_gold_daily",           "has_data": "Có",   "from_year": 2018, "notes": "Mua/bán, VND/chỉ, 2x/ngày"},
    {"category": "commodity", "indicator": "Giá bạc Phú Quý",                      "table": "vn_macro_silver_daily",         "has_data": "Có",   "from_year": 2020, "notes": "Mua/bán, VND"},
    {"category": "macro",     "indicator": "CPI Việt Nam MoM & YoY",               "table": "vn_gso_cpi_monthly",            "has_data": "Có",   "from_year": 2000, "notes": "GSO, monthly, overall + categories"},
    {"category": "macro",     "indicator": "PPI theo ngành",                        "table": "vn_gso_ppi_monthly",            "has_data": "Chưa", "from_year": None, "notes": "Schema ready, chưa crawl"},
    {"category": "macro",     "indicator": "GDP hàng quý theo ngành",               "table": "vn_gso_gdp_quarterly",          "has_data": "Chưa", "from_year": None, "notes": "Schema ready, chưa crawl"},
    {"category": "macro",     "indicator": "Xuất nhập khẩu hàng tháng",             "table": "vn_gso_trade_monthly",          "has_data": "Chưa", "from_year": None, "notes": "Schema ready, chưa crawl"},
    {"category": "macro",     "indicator": "Chỉ số IIP theo ngành",                 "table": "vn_gso_iip_monthly",            "has_data": "Chưa", "from_year": None, "notes": "Schema ready, chưa crawl"},
    {"category": "stock",     "indicator": "Giá OHLCV VN30 hàng ngày",             "table": "vn30_ohlcv_daily",              "has_data": "Có",   "from_year": 2018, "notes": "30 cổ phiếu HSX"},
    {"category": "stock",     "indicator": "Hồ sơ công ty VN30",                   "table": "vn30_company_profile",          "has_data": "Có",   "from_year": 2024, "notes": "ICB sector, vốn hóa"},
    {"category": "stock",     "indicator": "Kết quả kinh doanh hàng quý",           "table": "vn30_income_stmt_quarterly",    "has_data": "Có",   "from_year": 2019, "notes": "Revenue, net income, EPS"},
    {"category": "stock",     "indicator": "Bảng cân đối kế toán",                 "table": "vn30_balance_sheet_quarterly",  "has_data": "Có",   "from_year": 2019, "notes": "Assets, liabilities, equity"},
    {"category": "stock",     "indicator": "Lưu chuyển tiền tệ",                   "table": "vn30_cashflow_quarterly",       "has_data": "Có",   "from_year": 2019, "notes": "CFO, CFI, CFF"},
    {"category": "stock",     "indicator": "Chỉ số tài chính hàng ngày",            "table": "vn30_ratio_daily",              "has_data": "Có",   "from_year": 2020, "notes": "P/E, P/B, ROE, ROA, EPS"},
    {"category": "sentiment", "indicator": "News sentiment (future)",               "table": "—",                             "has_data": "Chưa", "from_year": None, "notes": "Planned — NLP on vnexpress, cafef"},
    {"category": "sentiment", "indicator": "Retail investor flow (future)",         "table": "—",                             "has_data": "Chưa", "from_year": None, "notes": "Planned — proprietary / exchange data"},
]

# ─────────────────────────────────────────────
# STYLING HELPERS
# ─────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
BORDER_SIDE = Side(border_style="thin", color="BFBFBF")
THIN_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
YES_FILL    = PatternFill("solid", fgColor="C6EFCE")
NO_FILL     = PatternFill("solid", fgColor="FFEB9C")
MAND_FILL   = PatternFill("solid", fgColor="E8F0FE")   # light blue for mandatory rows


def style_header(ws, row, columns):
    for col_idx, (header, width) in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def style_cell(cell, fill=None, bold=False):
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    if bold:
        cell.font = Font(bold=True)


# ─────────────────────────────────────────────
# SHEET BUILDERS
# ─────────────────────────────────────────────

def build_group_legend(wb):
    """Sheet 1: Group taxonomy legend + overview."""
    ws = wb.create_sheet("Groups & Overview", 0)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 20

    # Title
    ws.merge_cells("A1:C1")
    tc = ws.cell(row=1, column=1, value="Viet Dataverse — Data Group Taxonomy")
    tc.font = Font(bold=True, size=13, color="1F3864")
    tc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    ws.cell(row=2, column=1, value="Generated").font = Font(italic=True)
    ws.cell(row=2, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    # Group table header
    ws.row_dimensions[4].height = 22
    for col, (txt, w) in enumerate([("Group", 16), ("Description", 55), ("Tables Count", 14)], 1):
        c = ws.cell(row=4, column=col, value=txt)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, (gname, gdesc) in enumerate(GROUPS.items(), 5):
        count = sum(1 for t in DB_INVENTORY if t["group"] == gname)
        fill = PatternFill("solid", fgColor=GROUP_COLORS[gname])
        for col, val in enumerate([gname, gdesc, count or "—"], 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = fill
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center")
            if col == 1:
                c.font = Font(bold=True)

    # Mandatory columns section
    ws.row_dimensions[11].height = 22
    ws.merge_cells("A11:C11")
    mc = ws.cell(row=11, column=1, value="Mandatory Columns — Required in ALL tables")
    mc.font = Font(bold=True, size=11, color="1F3864")
    mc.fill = PatternFill("solid", fgColor="D9E1F2")
    mc.border = THIN_BORDER

    mandatory_info = [
        ("id",         "SERIAL PRIMARY KEY",       "Auto-increment, never manual MAX(id)+1"),
        ("period",     "DATE or VARCHAR(7)",        "YYYY-MM-DD (daily) or YYYY-MM (monthly)"),
        ("crawl_time", "TIMESTAMP NOT NULL",        "UTC timestamp when data was fetched"),
        ("source",     "TEXT",                      "Source URL or organisation name"),
        ("group_name", "VARCHAR(20) NOT NULL",      "Data group: macro | finance | commodity | stock | sentiment"),
    ]
    for col, hdr in enumerate(["Column", "Type", "Description"], 1):
        c = ws.cell(row=12, column=col, value=hdr)
        c.font = HEADER_FONT
        c.fill = PatternFill("solid", fgColor="2E5090")
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")

    for i, (col, typ, desc) in enumerate(mandatory_info, 13):
        fill = MAND_FILL
        for j, val in enumerate([col, typ, desc], 1):
            c = ws.cell(row=i, column=j, value=val)
            c.fill = fill
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center")
            if j == 1:
                c.font = Font(bold=True, color="1A5276")


def build_db_inventory(wb):
    ws = wb.create_sheet("System DB Inventory")
    ws.row_dimensions[1].height = 35

    columns = [
        ("Database", 22), ("Table Name", 32), ("Group", 12), ("Crawler", 30),
        ("Entity Key", 24), ("Period Column", 22), ("Frequency", 12),
        ("Source Org", 18), ("Description", 45), ("From Year", 10),
        ("~Rows", 10), ("Active", 8),
    ]
    style_header(ws, 1, columns)
    ws.freeze_panes = "A2"

    NO_CRAWLER_FILL = PatternFill("solid", fgColor="FFD7D7")

    for i, row in enumerate(DB_INVENTORY, 2):
        group = row["group"]
        g_fill = PatternFill("solid", fgColor=GROUP_COLORS.get(group, "FFFFFF"))
        crawler = row.get("crawler", "—")
        values = [
            row["db"], row["table"], row["group"], crawler,
            row["entity_key"], row["period_col"], row["freq"],
            row["source_org"], row["description"], row["from_year"],
            row["rows_approx"], "✅" if row["active"] else "⏸️",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            if col_idx == 3:  # group column colored
                cell.fill = g_fill
                cell.font = Font(bold=True)
            elif col_idx == 4 and str(val).startswith("⚠️"):
                cell.fill = NO_CRAWLER_FILL
                cell.font = Font(color="C0392B")
        ws.row_dimensions[i].height = 22

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(DB_INVENTORY)+1}"


def build_data_dictionary(wb):
    ws = wb.create_sheet("Data Dictionary")
    ws.row_dimensions[1].height = 40

    columns = [
        ("Table", 30), ("Mandatory", 10), ("Column", 20), ("Type", 22),
        ("Description", 50), ("Unit", 18), ("Example", 22),
    ]
    style_header(ws, 1, columns)
    ws.freeze_panes = "A2"

    all_rows = build_full_dictionary()
    prev_table = None

    for i, row in enumerate(all_rows, 2):
        is_mandatory = row.get("mandatory") == "✅"
        is_new_table = row["table"] != prev_table
        prev_table = row["table"]

        # Find group for coloring
        tbl_info = next((t for t in DB_INVENTORY if t["table"] == row["table"]), None)
        group = tbl_info["group"] if tbl_info else "macro"
        g_fill = PatternFill("solid", fgColor=GROUP_COLORS.get(group, "FFFFFF"))

        if is_mandatory:
            row_fill = MAND_FILL
        elif i % 2 == 0:
            row_fill = PatternFill("solid", fgColor="F5F5F5")
        else:
            row_fill = None

        values = [
            row["table"], row.get("mandatory", "—"), row["column"],
            row["type"], row["description"], row["unit"], row["example"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill
            if col_idx == 1 and is_new_table:
                cell.fill = g_fill
                cell.font = Font(bold=True)
            if col_idx == 2 and is_mandatory:
                cell.font = Font(bold=True, color="1A5276")
        ws.row_dimensions[i].height = 20

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(all_rows)+1}"


def build_stresstest_tracking(wb):
    ws = wb.create_sheet("Stresstest Inventory")
    ws.row_dimensions[1].height = 35

    columns = [
        ("Group", 12), ("Indicator", 38), ("Table Name", 32),
        ("Has Data?", 10), ("From Year", 10), ("Notes", 40),
    ]
    style_header(ws, 1, columns)
    ws.freeze_panes = "A2"

    for i, row in enumerate(STRESSTEST_TRACKING, 2):
        group = row["category"]
        g_fill = PatternFill("solid", fgColor=GROUP_COLORS.get(group, "FFFFFF"))
        values = [
            row["category"], row["indicator"], row["table"],
            row["has_data"], row["from_year"] or "—", row["notes"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            if col_idx == 1:
                cell.fill = g_fill
                cell.font = Font(bold=True)
            elif col_idx == 4:
                if val == "Có":
                    cell.fill = YES_FILL
                elif val == "Chưa":
                    cell.fill = NO_FILL
        ws.row_dimensions[i].height = 22

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(STRESSTEST_TRACKING)+1}"


# ─────────────────────────────────────────────
# MIGRATION STATUS DATA
# ─────────────────────────────────────────────

MIGRATION_SPEC = [
    # ── CRAWLING_BOT_DB ────────────────────────────────────────────────────────
    {
        "db": "CRAWLING_BOT_DB", "table": "vn_macro_sbv_rate_daily",
        "group": "finance", "source_fill": "sbv.gov.vn",
        "source_note": "Backfill WHERE source IS NULL",
        "group_fill": "finance",
    },
    {
        "db": "CRAWLING_BOT_DB", "table": "vn_macro_termdepo_daily",
        "group": "finance", "source_fill": "acb.com.vn",
        "source_note": "Backfill WHERE source IS NULL",
        "group_fill": "finance",
    },
    {
        "db": "CRAWLING_BOT_DB", "table": "vn_macro_fxrate_daily",
        "group": "finance", "source_fill": "(keep existing 'Crawl'/'API')",
        "source_note": "source already populated — skip",
        "group_fill": "finance",
    },
    {
        "db": "CRAWLING_BOT_DB", "table": "vn_macro_gold_daily",
        "group": "commodity", "source_fill": "api.btmc.vn (BTMC*) / 24h.com.vn (others)",
        "source_note": "CASE WHEN type LIKE 'BTMC%'",
        "group_fill": "commodity",
    },
    {
        "db": "CRAWLING_BOT_DB", "table": "vn_macro_silver_daily",
        "group": "commodity", "source_fill": "phuquy.com.vn",
        "source_note": "Backfill WHERE source IS NULL",
        "group_fill": "commodity",
    },
    {
        "db": "CRAWLING_BOT_DB", "table": "vn_gso_cpi_monthly",
        "group": "macro", "source_fill": "gso.gov.vn",
        "source_note": "Backfill WHERE source IS NULL",
        "group_fill": "macro",
    },
    {
        "db": "CRAWLING_BOT_DB", "table": "vn_gso_ppi_monthly",
        "group": "macro", "source_fill": "gso.gov.vn",
        "source_note": "Backfill WHERE source IS NULL",
        "group_fill": "macro",
    },
    {
        "db": "CRAWLING_BOT_DB", "table": "vn_gso_gdp_quarterly",
        "group": "macro", "source_fill": "gso.gov.vn",
        "source_note": "Backfill WHERE source IS NULL",
        "group_fill": "macro",
    },
    {
        "db": "CRAWLING_BOT_DB", "table": "vn_gso_trade_monthly",
        "group": "macro", "source_fill": "gso.gov.vn",
        "source_note": "Backfill WHERE source IS NULL",
        "group_fill": "macro",
    },
    {
        "db": "CRAWLING_BOT_DB", "table": "vn_gso_iip_monthly",
        "group": "macro", "source_fill": "gso.gov.vn",
        "source_note": "Backfill WHERE source IS NULL",
        "group_fill": "macro",
    },
    # ── GLOBAL_INDICATOR_DB ────────────────────────────────────────────────────
    {
        "db": "GLOBAL_INDICATOR_DB", "table": "global_macro",
        "group": "commodity", "source_fill": "Yahoo Finance",
        "source_note": "Already populated (crawler sets source)",
        "group_fill": "commodity",
    },
    # ── CRAWLING_CORP_DB ───────────────────────────────────────────────────────
    {
        "db": "CRAWLING_CORP_DB", "table": "vn30_ohlcv_daily",
        "group": "stock", "source_fill": "vnstock3",
        "source_note": "Backfill WHERE source IS NULL",
        "group_fill": "stock",
    },
    {
        "db": "CRAWLING_CORP_DB", "table": "vn30_company_profile",
        "group": "stock", "source_fill": "vnstock3",
        "source_note": "Backfill WHERE source IS NULL",
        "group_fill": "stock",
    },
    {
        "db": "CRAWLING_CORP_DB", "table": "vn30_income_stmt_quarterly",
        "group": "stock", "source_fill": "vnstock3",
        "source_note": "Backfill WHERE source IS NULL",
        "group_fill": "stock",
    },
    {
        "db": "CRAWLING_CORP_DB", "table": "vn30_balance_sheet_quarterly",
        "group": "stock", "source_fill": "vnstock3",
        "source_note": "Backfill WHERE source IS NULL",
        "group_fill": "stock",
    },
    {
        "db": "CRAWLING_CORP_DB", "table": "vn30_cashflow_quarterly",
        "group": "stock", "source_fill": "vnstock3",
        "source_note": "Backfill WHERE source IS NULL",
        "group_fill": "stock",
    },
    {
        "db": "CRAWLING_CORP_DB", "table": "vn30_ratio_daily",
        "group": "stock", "source_fill": "vnstock3",
        "source_note": "Backfill WHERE source IS NULL",
        "group_fill": "stock",
    },
]


def build_migration_status(wb):
    ws = wb.create_sheet("Migration Status")
    ws.row_dimensions[1].height = 35

    columns = [
        ("DB",          20),
        ("Table",       34),
        ("Group",       12),
        ("source — Backfill Value", 38),
        ("source — Logic",          34),
        ("group_name — Value",      18),
        ("Migration Script",        30),
        ("Status",                  14),
    ]
    style_header(ws, 1, columns)
    ws.freeze_panes = "A2"

    MIGRATED_FILL  = PatternFill("solid", fgColor="C6EFCE")  # green
    PENDING_FILL   = PatternFill("solid", fgColor="FFEB9C")  # yellow
    SCRIPT_FILL    = PatternFill("solid", fgColor="DDEEFF")

    for i, spec in enumerate(MIGRATION_SPEC, 2):
        group = spec["group"]
        g_fill = PatternFill("solid", fgColor=GROUP_COLORS.get(group, "FFFFFF"))

        values = [
            spec["db"],
            spec["table"],
            spec["group"],
            spec["source_fill"],
            spec["source_note"],
            spec["group_fill"],
            "migrate_mandatory_fields.py",
            "Pending",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            if col_idx == 3:
                cell.fill = g_fill
                cell.font = Font(bold=True)
            elif col_idx == 7:
                cell.fill = SCRIPT_FILL
                cell.font = Font(italic=True, color="1A5276")
            elif col_idx == 8:
                cell.fill = PENDING_FILL
                cell.font = Font(bold=True, color="7D6608")
        ws.row_dimensions[i].height = 22

    # ── Instructions block ──────────────────────────────────────────────────
    note_row = len(MIGRATION_SPEC) + 3
    ws.cell(row=note_row, column=1,
            value="▶ Run migration:").font = Font(bold=True, color="1A5276")
    ws.cell(row=note_row + 1, column=1,
            value="  python crawl_tools/migrate_mandatory_fields.py").font = Font(
                name="Courier New", size=9, color="1A5276")
    ws.cell(row=note_row + 2, column=1,
            value="  After run: update Status → ✅ Migrated and re-generate this file.").font = Font(
                italic=True, size=9, color="555555")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(MIGRATION_SPEC)+1}"


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_group_legend(wb)
    build_db_inventory(wb)
    build_data_dictionary(wb)
    build_stresstest_tracking(wb)
    build_migration_status(wb)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)

    all_dict_rows = build_full_dictionary()
    mandatory_count = sum(1 for r in all_dict_rows if r.get("mandatory") == "✅")
    specific_count  = len(all_dict_rows) - mandatory_count

    print(f"✅ Generated: {OUTPUT_PATH}")
    print(f"   Tables documented : {len(DB_INVENTORY)}")
    print(f"   Mandatory rows    : {mandatory_count}  ({len(DB_INVENTORY)} tables × 5 fields)")
    print(f"   Table-specific    : {specific_count} columns")
    print(f"   Indicators tracked: {len(STRESSTEST_TRACKING)}")
    print(f"   Migration entries : {len(MIGRATION_SPEC)} tables")
    print(f"   Sheets: Groups & Overview | System DB Inventory | Data Dictionary | Stresstest Inventory | Migration Status")


if __name__ == "__main__":
    main()
